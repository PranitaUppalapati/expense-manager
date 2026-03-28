"""SBI Savings Account — XLSX parser."""

from datetime import datetime
import openpyxl
from .base import AbstractBankParser, Transaction


class SBIParser(AbstractBankParser):
    BANK      = "SBI"
    BANK_ID   = "sbi"
    ACCT_TYPE = "debit"
    CURRENCY  = "INR"

    def parse(self, filepath: str) -> list[Transaction]:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Prefer "All Transactions" sheet; fall back to active sheet
        ws = wb["All Transactions"] if "All Transactions" in wb.sheetnames else wb.active

        # Extract masked account number from header rows (row 2)
        account = "XXXXXXX"
        for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
            for cell in row:
                if cell and "XXXXXXX" in str(cell):
                    import re
                    m = re.search(r"XXXXXXX\d+", str(cell))
                    if m:
                        account = m.group()
                    break

        txns = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            _period, date, desc, _ref, credit, debit, balance, _src = row[:8]
            if not date or (not credit and not debit):
                continue
            txns.append(Transaction(
                bank=self.BANK,
                bank_id=self.BANK_ID,
                account=account,
                account_type=self.ACCT_TYPE,
                currency=self.CURRENCY,
                date=self._parse_date(date),
                description=(desc or "").strip(),
                credit=float(credit) if credit else 0.0,
                debit=float(debit)  if debit  else 0.0,
                balance=float(balance) if balance else 0.0,
            ))
        return txns

    @staticmethod
    def _parse_date(raw) -> str:
        for fmt in ("%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(str(raw), fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        return str(raw)
